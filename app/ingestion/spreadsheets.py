"""Spreadsheet Ingestion (Phase 11.4).

Ingests Excel (.xlsx, .xls) and CSV files, normalizing sheets/cells
while retaining workbook provenance and cell-level traceability.
"""

from __future__ import annotations

import csv
import hashlib
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from uuid import UUID

import openpyxl
import pandas as pd

from app.config import get_settings
from app.ingestion.chunking import TextSegment
from app.ingestion.multimodal import MultimodalType, Spreadsheet, SpreadsheetCell, SpreadsheetSheet
from app.logging_config import get_logger

logger = get_logger("argus.ingestion.spreadsheets")


@dataclass(frozen=True)
class SpreadsheetResult:
    """Result of spreadsheet ingestion."""
    sheets: list[SpreadsheetSheet]
    author: str | None
    created_date: datetime | None
    modified_date: datetime | None
    total_rows: int
    total_cells: int
    metadata: dict[str, Any]


def _read_csv_file(file_path: Path) -> list[list[Any]]:
    """Read CSV file and return as list of rows."""
    rows = []
    with open(file_path, "r", encoding="utf-8", newline="") as f:
        sample = f.read(1024)
        f.seek(0)

        delimiter = ","
        if sample.strip():
            try:
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
            except csv.Error:
                pass

        reader = csv.reader(f, delimiter=delimiter)
        rows = list(reader)
        return rows


def _read_excel_file(file_path: Path) -> dict[str, list[list[Any]]]:
    """Read Excel file and return dict of sheet_name -> rows."""
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    sheets = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            # Stop at completely empty rows at the end
            if all(cell is None for cell in row):
                break
            rows.append(list(row))
        sheets[sheet_name] = rows
    
    workbook.close()
    return sheets


def _rows_to_sheet(
    sheet_name: str,
    rows: list[list[Any]],
    file_path: Path,
) -> SpreadsheetSheet:
    """Convert raw rows to SpreadsheetSheet with cell-level provenance."""
    cells = []
    data_rows = rows[1:] if rows else []  # Assume first row is header
    headers = rows[0] if rows else []
    
    for row_idx, row in enumerate(data_rows):
        for col_idx, cell_value in enumerate(row):
            if cell_value is None:
                continue
            
            # Convert value to string for storage
            str_value = str(cell_value) if cell_value is not None else ""
            if not str_value.strip():
                continue
            
            # Check if cell has a formula (only for openpyxl)
            formula = None
            # Note: With data_only=True, formulas are evaluated, not preserved
            
            cell = SpreadsheetCell(
                sheet=sheet_name,
                row=row_idx,  # 0-based in data rows
                col=col_idx,
                value=str_value,
                formula=formula,
                data_type=type(cell_value).__name__ if cell_value is not None else None,
            )
            cells.append(cell)
    
    return SpreadsheetSheet(
        name=sheet_name,
        rows=len(data_rows),
        cols=len(headers) if headers else (max(len(r) for r in rows) if rows else 0),
        cells=cells,
    )


def ingest_spreadsheet(file_path: Path) -> SpreadsheetResult:
    """Ingest a spreadsheet file (Excel or CSV) with full provenance.
    
    Returns SpreadsheetResult with normalized sheets, cells, and metadata.
    """
    settings = get_settings()
    
    if not settings.multimodal_enabled or not settings.multimodal_spreadsheet_enabled:
        raise RuntimeError("Spreadsheet ingestion disabled via configuration")
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {file_path}")
    
    suffix = file_path.suffix.lower()
    sheets = []
    metadata = {
        "filename": file_path.name,
        "size_bytes": file_path.stat().st_size,
        "format": suffix,
    }
    
    total_rows = 0
    total_cells = 0
    author = None
    created_date = None
    modified_date = None
    
    if suffix == ".csv":
        rows = _read_csv_file(file_path)
        if rows:
            sheet = _rows_to_sheet("Sheet1", rows, file_path)
            sheets.append(sheet)
            total_rows = sheet.rows
            total_cells = len(sheet.cells)
    
    elif suffix in (".xlsx", ".xls", ".xlsm"):
        # Read with openpyxl for metadata
        wb = openpyxl.load_workbook(file_path, read_only=True)
        author = wb.properties.creator
        created_date = wb.properties.created
        modified_date = wb.properties.modified
        
        # Read data with pandas for better handling
        xl = pd.ExcelFile(file_path)
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
            rows = df.values.tolist()
            if rows:
                sheet = _rows_to_sheet(sheet_name, rows, file_path)
                sheets.append(sheet)
                total_rows += sheet.rows
                total_cells += len(sheet.cells)
        wb.close()
    
    else:
        raise ValueError(f"Unsupported spreadsheet format: {suffix}")
    
    if not sheets:
        raise ValueError(f"No data found in spreadsheet: {file_path}")
    
    metadata.update({
        "sheet_count": len(sheets),
        "sheet_names": [s.name for s in sheets],
    })
    
    return SpreadsheetResult(
        sheets=sheets,
        author=author,
        created_date=created_date,
        modified_date=modified_date,
        total_rows=total_rows,
        total_cells=total_cells,
        metadata=metadata,
    )


def spreadsheet_to_multimodal(
    result: SpreadsheetResult,
    document_id: UUID,
    source_path: str,
    source_chunk_ids: list[UUID] | None = None,
) -> Spreadsheet:
    """Convert SpreadsheetResult to Multimodal Spreadsheet object for storage."""
    return Spreadsheet(
        id=UUID(int=0),  # Will be assigned by store
        source_path=source_path,
        content_type=MultimodalType.SPREADSHEET,
        source_chunk_ids=source_chunk_ids or [],
        sheets=result.sheets,
        author=result.author,
        created_date=result.created_date,
        modified_date=result.modified_date,
        metadata=result.metadata,
    )


def spreadsheet_to_text_segments(result: SpreadsheetResult) -> list[TextSegment]:
    """Convert spreadsheet content to text segments for chunking pipeline."""
    segments = []
    
    for sheet in result.sheets:
        if not sheet.cells:
            continue
        
        # Build text representation of sheet
        lines = [f"Sheet: {sheet.name}"]
        
        # Group cells by row
        rows_dict: dict[int, dict[int, object]] = {}
        for cell in sheet.cells:
            if cell.row not in rows_dict:
                rows_dict[cell.row] = {}
            rows_dict[cell.row][cell.col] = cell.value
        
        # Build rows
        for row_idx in sorted(rows_dict.keys()):
            row_data = rows_dict[row_idx]
            if not row_data:
                continue
            max_col = max(row_data.keys()) if row_data else 0
            row_values = [row_data.get(c, "") for c in range(max_col + 1)]
            lines.append(" | ".join(str(v) for v in row_values))
        
        sheet_text = "\n".join(lines)
        
        if sheet_text.strip():
            segments.append(TextSegment(
                text=sheet_text,
                page_start=1,
                page_end=1,
                char_start=None,
                char_end=None,
                section_path=f"Sheet: {sheet.name}",
                metadata={
                    "multimodal_type": "spreadsheet",
                    "sheet": sheet.name,
                    "rows": sheet.rows,
                    "cols": sheet.cols,
                    "cell_count": len(sheet.cells),
                },
            ))
    
    return segments


def compute_spreadsheet_checksum(file_path: Path) -> str:
    """Compute checksum for spreadsheet file."""
    content = file_path.read_bytes()
    return hashlib.sha256(content).hexdigest()


def is_valid_spreadsheet(file_path: Path) -> bool:
    """Validate if file is a supported spreadsheet format."""
    return file_path.suffix.lower() in (".csv", ".xlsx", ".xls", ".xlsm")


__all__ = [
    "SpreadsheetResult",
    "compute_spreadsheet_checksum",
    "ingest_spreadsheet",
    "is_valid_spreadsheet",
    "spreadsheet_to_multimodal",
    "spreadsheet_to_text_segments",
]